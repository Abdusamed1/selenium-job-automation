import pytest
from openpyxl import load_workbook
from pathlib import Path
from Code.job_scrapper import save_to_excel  # Update this with the actual module name

# Sample jobs data
jobs = [
    {'title': 'Software Engineer', 'location': 'New York', 'url': 'http://example.com'},
    {'title': 'Data Scientist', 'location': 'San Francisco', 'url': 'http://example.com/2'}
]

@pytest.fixture
def temp_excel_file(tmp_path):
    """Fixture to create a temporary Excel file for testing."""
    file = tmp_path / "job_listings.xlsx"
    return file

def test_create_new_file_no_duplicates(temp_excel_file):
    """Test that a new file is created when it doesn't exist, with no duplicates."""
    # Call the function to create a new Excel file
    save_to_excel(None, jobs, filename=temp_excel_file)

    wb = load_workbook(temp_excel_file)
    ws = wb.active

    assert ws['A1'].value == 'Job Title'
    assert ws['B1'].value == 'Location'
    assert ws['C1'].value == 'URL'

    assert ws['A2'].value == 'Software Engineer'
    assert ws['B2'].value == 'New York'
    assert ws['C2'].value == 'http://example.com'

    assert ws['A3'].value == 'Data Scientist'
    assert ws['B3'].value == 'San Francisco'
    assert ws['C3'].value == 'http://example.com/2'


def test_append_without_duplicates(temp_excel_file):
    save_to_excel(None, [jobs[0]], filename=temp_excel_file)
    save_to_excel(None, jobs, filename=temp_excel_file)
    wb = load_workbook(temp_excel_file)
    ws = wb.active

    assert ws['A1'].value == 'Job Title'
    assert ws['B1'].value == 'Location'
    assert ws['C1'].value == 'URL'

    assert ws['A2'].value == 'Software Engineer'
    assert ws['B2'].value == 'New York'
    assert ws['C2'].value == 'http://example.com'

    assert ws['A3'].value == 'Data Scientist'
    assert ws['B3'].value == 'San Francisco'
    assert ws['C3'].value == 'http://example.com/2'

    assert ws.max_row == 3
