from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import undetected_chromedriver as uc
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import date

class Indeed_Script():
    def __init__(self):
        self.options = Options()
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        # self.options.add_argument('--headless')
        self.options.add_argument(f'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36')
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.driver = uc.Chrome(options=self.options)
        self.wait = WebDriverWait(self.driver, 10)

    def random_sleep(self, min_seconds=2, max_seconds=5):
        time.sleep(random.uniform(min_seconds, max_seconds))

    def human_like_scroll(self):
        scroll_pause_time = random.uniform(0.5, 2.0)
        screen_height = self.driver.execute_script("return window.screen.height;")
        for i in range(1, random.randint(2,5)):
            self.driver.execute_script(f"window.scrollTo(0, {screen_height}*{i});")
            time.sleep(scroll_pause_time)

    def indeed_script(self):
        try:
            self.driver.get("https://www.indeed.com/q-usa-jobs.html")
            # self.driver.get("https://www.indeed.ca/q-can-jobs.html")

            time.sleep(10)
            self.driver.find_element(By.XPATH, "//*[@id='text-input-what']").click()
            self.random_sleep()

            self.driver.find_element(By.XPATH,
                                     "//*[@id='jobsearch']/div/div[1]/div[1]/div/div/span/span[2]/button").click()
            self.random_sleep()

            # TYPE JOB TITLE IN THE QUTATION MARKS BELOW
            jobfield = self.wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='text-input-what']")))
            jobfield.send_keys("")

            location = self.wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='text-input-where']")))
            location.send_keys("United States")

            search = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[@type='submit' and contains(text(), 'Search')]")))
            search.click()

            time.sleep(2)
            dateposted = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[@id='filter-dateposted']")))
            dateposted.click()

            # ONLY CHANGE 'Last 24 hours' TO EITHER 'Last 3 days' OR 'Last 7 days' OR 'Last 14 days'
            time.sleep(2)
            lastdayjobs = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[text()= 'Last 24 hours']")))
            lastdayjobs.click()
            time.sleep(3)
            self.random_sleep()

            experience = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='filter-explvl']")))
            experience.click()

            # FOR ENTRY LEVEL KEEP 'Enter Level'. FOR MID LEVEL USE 'Mid Level'. FOR SENIOR LEVEL USE 'Senior Level'
            self.random_sleep()
            entry_level = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[text()='Entry Level']")))
            entry_level.click()
            self.random_sleep()

        except Exception as e:
            print(f"An error occurred: {e}")

    def job_info(self):
        try:
            listofjobs = []
            job_cards = self.driver.find_elements(By.XPATH, "//*[@id='mosaic-provider-jobcards']/ul/li/div")
            next_page = self.driver.find_elements(By.XPATH, "//*[@id='jobsearch-JapanPage']/div/div[5]/div/div[1]/nav/ul/li/a")
            num_cards = len(job_cards)
            # next_pagelength = len(next_page) + 1

            print(f"Number of job cards found: {num_cards}")
            # print(f"Number of pages found: {next_pagelength}")

            # for j in range(1, next_pagelength + 1):
            #     time.sleep(6)
            for i in range(1, num_cards + 1):
                try:
                    job_card = self.wait.until(EC.visibility_of_element_located(
                        (By.XPATH, f"//*[@id='mosaic-provider-jobcards']/ul/li[{i}]/div")))
                    if "Strengthen your profile" in job_card.text:
                        continue

                    job_card.click()
                    time.sleep(2)
                    job_title = self.wait.until(EC.visibility_of_element_located((By.XPATH,
                                                                                          "//*[@id='jobsearch-ViewjobPaneWrapper']/div/div[2]/div[2]/div[1]/div/div[1]/div[1]/h2/span"))).text
                    time.sleep(2)
                    self.random_sleep()
                    job_location = self.wait.until(EC.visibility_of_element_located((By.XPATH,
                                                                                             "//*[@id='jobsearch-ViewjobPaneWrapper']/div/div[2]/div[2]/div[1]/div/div[1]/div[2]/div/div/div/div[2]/div"))).text
                    time.sleep(2)
                    self.random_sleep()
                    job_url = self.driver.current_url

                    listofjobs.append({'title': job_title, 'location': job_location, 'url': job_url})
                    print(f"Job Title: {job_title}, Location: {job_location}, URL: {job_url}")
                    self.random_sleep()

                except Exception as e:
                    print(f"An error occurred for job card {i}: {e}")
                # try:
                #     j+=1
                #     clicknextpage=  self.driver.find_element(By.XPATH, f"//*[@id='jobsearch-JapanPage']/div/div[5]/div/div[1]/nav/ul/li[{j}]/a")
                #     time.sleep(2)
                #     clicknextpage.click()
                #     time.sleep(5)
                #     self.random_sleep()
                # except Exception as e:
                #     print(f"No more pages or failed to find next button: {e}")
                #     break

            print("List of jobs:", listofjobs)
            return listofjobs

        except Exception as e:
            print(f"An error occurred: {e}")

        finally:
            self.driver.quit()

    def save_to_excel(self, jobs, filename='job_listings.xlsx'):
        my_file = Path(filename)

        if my_file.is_file():
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Job Title', 'Location', 'URL'])

        if ws['O1'].value is None:
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 105
            ws.column_dimensions["D"].width = 10

            ws['D1'] = 'Date'
            ws.column_dimensions["E"].width = 10
            ws['E1'] = 'Applied'
            ws.column_dimensions["F"].width = 15
            ws['F1'] = 'Offer/Rejected'

        existing_urls = set(row[2] for row in ws.iter_rows(min_row=2, values_only=True))

        for job in jobs:
            if job['location'] and job['url'] not in existing_urls:
                ws.append([job['title'], job['location'], job['url'], date.today()])

        wb.save(filename)


if __name__ == "__main__":
    obj = Indeed_Script()
    obj.indeed_script()
    job_list = obj.job_info()
    if job_list:
        obj.save_to_excel(job_list)

